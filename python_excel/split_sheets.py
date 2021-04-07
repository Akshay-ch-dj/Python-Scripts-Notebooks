import os
from openpyxl import load_workbook, Workbook

path = (
    "/home/akshay/Downloads/files_to_split/"  # place path where files to split up are
)

targetdir = path + "splitted_sheets/"  # where you want your new files

if not os.path.exists(targetdir):  # makes your new directory if not exists
    os.makedirs(targetdir)

for root, dir, files in os.walk(path, topdown=False):  # all the files you want to split
    xlsxfiles = [f for f in files]  # can add selection condition here

for f in xlsxfiles:
    wb = load_workbook(os.path.join(root, f))

    for sheet in wb.worksheets:  # cycles through each sheet in each workbook
        new_wb = Workbook()  # makes a temp copy of that book
        ws = new_wb.active

        for row_data in sheet.iter_rows():
            for row_cell in row_data:
                ws[row_cell.coordinate].value = row_cell.value

        # saves each sheet as the sheet name
        new_wb.save("{0}.xlsx".format(targetdir + sheet.title))
